import io
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

# Configurazione Pagina Streamlit
st.set_page_config(
    page_title="Riclassificazione Fatture Trasporti", layout="wide"
)

st.title("🚚 Tool Riclassificazione Fatture Trasporti per Competenza")

# ---------------------------------------------------------
# SIDEBAR: Upload File Excel
# ---------------------------------------------------------
st.sidebar.header("📁 Upload Dati Input")

file_transazioni = st.sidebar.file_uploader(
    "1. Lista Transazioni (Excel)", type=["xlsx", "xls"]
)
file_fornitori = st.sidebar.file_uploader(
    "2. Lista Fornitori (Excel)", type=["xlsx", "xls"]
)


# ---------------------------------------------------------
# FUNZIONI DI ELABORAZIONE LOGICA
# ---------------------------------------------------------
def calcola_mese_calcolo(data):
    """Step 3: Determina il Mese Calcolo in base alla data di registrazione (Cutoff giorno 3).

    - Se giorno >= 3: Mese Corrente
    - Se giorno < 3: Mese Precedente
    """
    if pd.isna(data):
        return pd.NaT

    if data.day >= 3:
        return data.to_period("M")
    else:
        return (data - pd.DateOffset(months=1)).to_period("M")


def elabora_dati(df_trans, df_forn):
    # Standardizzazione nomi colonne
    df_trans.columns = df_trans.columns.str.strip()
    df_forn.columns = df_forn.columns.str.strip()

    # ---------------------------------------------------------
    # STEP 1: Estrazione Codice Fornitore e Lookup Nome
    # ---------------------------------------------------------
    df_trans["Codice_Fornitore_Estratto"] = (
        df_trans["Valore visualizzato conto"]
        .astype(str)
        .str.extract(r"--(\d+)$")
    )

    df_trans["Codice_Fornitore_Estratto"] = df_trans[
        "Codice_Fornitore_Estratto"
    ].str.strip()
    df_forn["Account fornitore"] = (
        df_forn["Account fornitore"].astype(str).str.strip()
    )

    # Merge con Anagrafica Fornitori per recuperare 'Nome'
    df_merged = df_trans.merge(
        df_forn[["Account fornitore", "Nome"]],
        left_on="Codice_Fornitore_Estratto",
        right_on="Account fornitore",
        how="left",
    )

    df_merged["Nome"] = df_merged["Nome"].fillna("FORNITORE NON TROVATO")

    # ---------------------------------------------------------
    # STEP 2 & 2.bis: Filtro Esclusione "Trasporti mm.yyyy" e "Trasporti mm.yy"
    # ---------------------------------------------------------
    pattern_esclusione = r"(?i)trasporti\s+\d{1,2}\.\d{2,4}"
    df_filtered = df_merged[
        ~df_merged["Descrizione"]
        .astype(str)
        .str.contains(pattern_esclusione, regex=True, na=False)
    ].copy()

    # ---------------------------------------------------------
    # STEP 3: Mese Calcolo
    # ---------------------------------------------------------
    df_filtered["Data"] = pd.to_datetime(df_filtered["Data"], errors="coerce")
    df_filtered["Mese Calcolo"] = df_filtered["Data"].apply(
        calcola_mese_calcolo
    )

    # ---------------------------------------------------------
    # STEP 4: Fix Data Doc Blank & Competenza da Data Doc
    # ---------------------------------------------------------
    df_filtered["Data documento"] = pd.to_datetime(
        df_filtered["Data documento"], errors="coerce"
    )

    # Se DATA DOC è vuota, utilizziamo la data di registrazione (DATA)
    df_filtered["Data documento"] = df_filtered["Data documento"].fillna(
        df_filtered["Data"]
    )

    # Competenza deriva unicamente dal mese della Data Documento
    df_filtered["Competenza"] = df_filtered["Data documento"].dt.to_period("M")

    # Estrazione numero del mese di competenza (1, 2, 3...) per layout simile all'immagine
    df_filtered["Mese_Competenza_Num"] = df_filtered["Data documento"].dt.month
    df_filtered["Mese_Calcolo_Num"] = df_filtered["Mese Calcolo"].dt.month

    # Formattazione in stringa YYYY-MM
    df_filtered["Mese Calcolo Str"] = df_filtered["Mese Calcolo"].astype(str)
    df_filtered["Competenza Str"] = df_filtered["Competenza"].astype(str)

    return df_filtered


def applica_stile_excel_pivot(ws, titolo_sezione="ITALIA"):
    """Applica la formattazione visuale identica allo screenshot Excel.

    - Riga 1: Banner Verde con Titolo
    - Riga 3/4: Header azzurro (#9FC5E8)
    - Formattazione numeri: #,##0 (interi)
    - Riga Finale Totali in azzurro
    """
    fill_verde = PatternFill(
        start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
    )
    fill_azzurro = PatternFill(
        start_color="9FC5E8", end_color="9FC5E8", fill_type="solid"
    )
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # Inserisci 3 righe in alto per far spazio all'intestazione
    ws.insert_rows(1, amount=3)

    # Titolo Sezione A1
    ws["A1"] = titolo_sezione
    ws["A1"].font = font_bold
    ws["A1"].fill = fill_verde
    ws["A1"].alignment = align_center

    # Label misura in A3 e B3
    ws["A3"] = "Sum of Importo nella valuta di dichiarazione"
    ws["A3"].font = font_bold
    ws["B3"] = "Column Labels"
    ws["B3"].font = font_bold

    max_col = ws.max_column
    max_row = ws.max_row

    # Formattazione Header Pivot (Riga 4)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = fill_azzurro
        cell.font = font_bold
        if col > 1:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

    # Formattazione Dati e Totali
    for r in range(5, max_row + 1):
        is_grand_total_row = r == max_row
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_bold if is_grand_total_row else font_regular

            if is_grand_total_row:
                cell.fill = fill_azzurro

            if c > 1:
                # Formato numero intero con separatore migliaia come in foto
                cell.number_format = "#,##0"
                cell.alignment = align_right
            else:
                cell.alignment = align_left

    # Regolazione automatica larghezza colonne
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def genera_excel_export(
    pivot_estero, pivot_nazionali, pivot_registrazione_comp, df_dettaglio
):
    """Genera un file Excel in memoria formattato ed esteticamente coerente con lo screenshot."""
    output = io.BytesIO()

    # Usiamo direttamente pd.ExcelWriter
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1. Scrittura delle 3 Pivot Table nei rispettivi Fogli
        pivot_estero.to_excel(writer, sheet_name="Estero")
        pivot_nazionali.to_excel(writer, sheet_name="Nazionali")
        pivot_registrazione_comp.to_excel(writer, sheet_name="Reg_vs_Competenza")

        # 2. Scrittura Dettaglio Transazioni
        df_export = df_dettaglio.copy()
        if "Data" in df_export.columns:
            df_export["Data"] = df_export["Data"].dt.strftime("%Y-%m-%d")
        if "Data documento" in df_export.columns:
            df_export["Data documento"] = df_export[
                "Data documento"
            ].dt.strftime("%Y-%m-%d")

        df_export.to_excel(writer, sheet_name="Dettaglio Transazioni", index=False)

        # 3. Accesso al workbook openpyxl dal writer per applicare lo stile grafico
        wb = writer.book

        applica_stile_excel_pivot(wb["Estero"], titolo_sezione="ESTERO")
        applica_stile_excel_pivot(wb["Nazionali"], titolo_sezione="ITALIA")
        applica_stile_excel_pivot(
            wb["Reg_vs_Competenza"],
            titolo_sezione="ANALISI REGISTRAZIONE VS COMPETENZA",
        )

        # Stile Intestazione Dettaglio Transazioni
        ws_dettaglio = wb["Dettaglio Transazioni"]
        header_fill = PatternFill(
            start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
        )
        for cell in ws_dettaglio[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

    return output.getvalue()


# ---------------------------------------------------------
# INTERFACCIA UTENTE & ISTRUZIONI INIZIALI
# ---------------------------------------------------------
if not (file_transazioni and file_fornitori):
    st.info(
        "📌 **Per il corretto funzionamento dell'applicazione, assicurati di scaricare e caricare i file richiesti seguendo queste indicazioni:**"
    )

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
        ### 📄 File Transazioni
        1. Andare in **Transazioni Giustificativo**.
        2. Impostare le seguenti colonne nella tabella:
           * `Numero giornale di registrazione`
           * `Documento`
           * `Giustificativo`
           * `Data`
           * `Data documento`
           * `Anno chiuso`
           * `Conto principale`
           * `Valore visualizzato conto`
           * `Nome conto`
           * `Descrizione`
           * `Valuta`
           * `Importo nella valuta della transazione`
           * `Importo`
           * `Importo nella valuta di dichiarazione`
        3. Filtrare la colonna **Conto Principale** con l'opzione **"è uno di"** copiando i codici di conto:
           * `07.02.02.060.02`
           * `07.02.02.060.03`
        4. Filtrare inoltre la data dal **1/1/2026**.
        5. Cliccare con tasto destro e selezionare **"Esporta tutte le righe"**. L'excel verrà poi scaricato.
        """)

    with col2:
        st.markdown("""
        ### 👤 File Anagrafica Fornitori
        1. Selezionare la voce **"Tutti i fornitori"**.
        2. Cliccare con tasto destro sulla colonna e selezionare **"Esporta tutte le righe"**.
        3. L'excel verrà poi scaricato.
        """)

else:
    # ---------------------------------------------------------
    # ESECUZIONE ELABORAZIONE
    # ---------------------------------------------------------
    try:
        df_trans = pd.read_excel(file_transazioni)
        df_forn = pd.read_excel(file_fornitori)

        with st.spinner("Elaborazione dati in corso..."):
            df_elaborato = elabora_dati(df_trans, df_forn)

        st.success("✅ Dati elaborati con successo!")

        # ---------------------------------------------------------
        # PIVOT TABLES (Usando i numeri del mese 1..12 per matchare lo screenshot)
        # ---------------------------------------------------------

        # 1. Estero (Conto 07.02.02.060.03 o TRASPORTI SU VENDITE ESTERO)
        df_estero = df_elaborato[
            df_elaborato["Nome conto"].astype(str).str.upper()
            == "TRASPORTI SU VENDITE ESTERO"
        ]

        pivot_estero = pd.pivot_table(
            df_estero,
            values="Importo nella valuta di dichiarazione",
            index=["Nome"],
            columns=["Mese_Competenza_Num"],
            aggfunc="sum",
            fill_value=0.0,
            margins=True,
            margins_name="Grand Total",
        )
        pivot_estero.index.name = "Row Labels"

        # 2. Nazionali (Conto 07.02.02.060.02 o TRASPORTI SU VENDITE NAZIONALI)
        df_nazionali = df_elaborato[
            df_elaborato["Nome conto"].astype(str).str.upper()
            == "TRASPORTI SU VENDITE NAZIONALI"
        ]

        pivot_nazionali = pd.pivot_table(
            df_nazionali,
            values="Importo nella valuta di dichiarazione",
            index=["Nome"],
            columns=["Mese_Competenza_Num"],
            aggfunc="sum",
            fill_value=0.0,
            margins=True,
            margins_name="Grand Total",
        )
        pivot_nazionali.index.name = "Row Labels"

        # 3. Data Transazione (Mese Calcolo) vs Data Competenza
        pivot_registrazione_comp = pd.pivot_table(
            df_elaborato,
            values="Importo nella valuta di dichiarazione",
            index=["Mese Calcolo Str"],
            columns=["Mese_Competenza_Num"],
            aggfunc="sum",
            fill_value=0.0,
            margins=True,
            margins_name="Grand Total",
        )
        pivot_registrazione_comp.index.name = "Mese Registrazione"

        # ---------------------------------------------------------
        # EXPORT EXCEL & DASHBOARD
        # ---------------------------------------------------------
        st.subheader("📥 Export Report Excel")
        excel_data = genera_excel_export(
            pivot_estero,
            pivot_nazionali,
            pivot_registrazione_comp,
            df_elaborato,
        )

        st.download_button(
            label="💾 Scarica Report Excel Formattato",
            data=excel_data,
            file_name="Riclassificazione_Trasporti_Competenza.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.markdown("---")

        # Visualizzazione Tab Dashboard
        tab1, tab2, tab3, tab4 = st.tabs(
            [
                "🇮🇹 Trasporti Nazionali",
                "🌍 Trasporti Estero",
                "🔄 Data Registrazione vs Competenza",
                "📋 Dettaglio Dati Processati",
            ]
        )

        with tab1:
            st.markdown("### Trasporti Su Vendite Nazionali (ITALIA)")
            st.dataframe(
                pivot_nazionali.style.format("{:,.0f}"),
                use_container_width=True,
            )

        with tab2:
            st.markdown("### Trasporti Su Vendite Estero")
            st.dataframe(
                pivot_estero.style.format("{:,.0f}"), use_container_width=True
            )

        with tab3:
            st.markdown("### Analisi Data Registrazione vs Competenza")
            st.dataframe(
                pivot_registrazione_comp.style.format("{:,.0f}"),
                use_container_width=True,
            )

        with tab4:
            st.markdown("### Transazioni Trasformate (Anteprima)")
            cols_preview = [
                "Documento",
                "Data",
                "Data documento",
                "Codice_Fornitore_Estratto",
                "Nome",
                "Nome conto",
                "Descrizione",
                "Importo nella valuta di dichiarazione",
                "Mese Calcolo Str",
                "Competenza Str",
            ]
            st.dataframe(
                df_elaborato[cols_preview], use_container_width=True
            )

    except Exception as e:
        st.error(f"Errore durante l'elaborazione dei file: {e}")